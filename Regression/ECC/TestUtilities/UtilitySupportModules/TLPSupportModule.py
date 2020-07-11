#A collection of TLP file mocking and NRG_regression data processing TLP_Enrollments_Electric, plus some global variables.

#Author: Matt Hissong

#NRG_regression Custom Modules
import GenericSettings

#Python Standard Library and Third Party Modules
import sys
import os
import random
import collections
from collections import OrderedDict
import openpyxl
#from sortedcontainers import SortedList

def utilityMaskSearch(utilityCode):
    myQueryString = """select top 1 MaskLength, MaskBase, MaskSQL, *
                        from DataServices%s.dbo.UtilityAccountMask
                        where AccountNumberType = 'utility'
                        and UtilityID = '%s'
                        and StartDate <= CURRENT_TIMESTAMP
					    order by StartDate desc, DataServices%s.dbo.UtilityAccountMask.InsertDT desc
                        """ % (GenericSettings.getMyEnvironment(), GenericSettings.stripLeadingZeroesFromString(utilityCode), GenericSettings.getMyEnvironment())
    result = GenericSettings.genericSQLQuery(myQueryString)
    backupTableSearchAuthorized = True
    if (result == None) or (len(result) < 1):
        print("\nThis is myQueryString: " + myQueryString + "\n")
        print("\n No rules for the UAN found in the Dataservices.dbo.UtilityAccountMask table.  Utility code is: " + utilityCode + "\n")
        ##print("\n No rules for the UAN found in the Dataservices.dbo.Utilities table.  This indeterminate result is judged as valid for the moment.\n")
        # return []
        if (backupTableSearchAuthorized):
            print("Trying the old table DataServices%s.dbo.Utilities" % GenericSettings.getMyEnvironment())
            myQueryString = """select AccountLength, AccountMask, AccountMaskSQL, *
                    from DataServices%s.dbo.Utilities
                    where UtilityID = '%s'
                    """ % (GenericSettings.getMyEnvironment(), GenericSettings.stripLeadingZeroesFromString(utilityCode))
            result = GenericSettings.genericSQLQuery(myQueryString)
            if (result == None) or (len(result) < 1):
                print("\nThis is myQueryString: " + myQueryString + "\n")
                sys.exit("\n No rules for the UAN found in any relevant table, including Dataservices.dbo.UtilityAccountMask.  Utility code is: " + utilityCode + "\n")
            elif (len(result) > 1):
                sys.exit("\nMore than one entry for the UtilityCode %s in DataServices%s.dbo.Utilities.  Aborting.\n" % (utilityCode, GenericSettings.getMyEnvironment() ) )
        else:
            sys.exit("Because relying on DataServices%s.dbo.Utilities (\"the backup table\") for mask data is disallowed by the backupTableSearchAuthorized flag, the program has to terminate here.\n" % GenericSettings.getMyEnvironment())
    return result

#replacement example
#result = utilityMaskSearch(utilityCode)

#takes an alphanumeric string with numbers and letters and increments it from right to left... if the last letter is a num, it ups that 'til
#it rolls over into the next num over.  Same with a letter, but in the letter space... and if there's rollover, it checks the next
#slot over and makes sure that one stays the same type that it was originally.  #This function ignores special characters and skips
#to the next slot over after them.
def incrStrNum(someStrNum, hardStopCharList):
    if(someStrNum is None):
        return someStrNum
    if(someStrNum == ""):
        return someStrNum
    #turn the string-num into a list of characters, because Python's strings are immutable.
    myStrNum = list(someStrNum)
    currentInd = len(myStrNum) - 1
    rollover = True
    while(rollover and (currentInd >= 0)):
        rollover = False
        myTempChar = str(myStrNum[currentInd])
        if(myTempChar in hardStopCharList):
            return "".join(myStrNum)
        elif(myTempChar == 'z'):
            myStrNum[currentInd] = 'a'
            rollover = True
        elif (myTempChar == 'Z'):
            myStrNum[currentInd] = 'A'
            rollover = True
        elif(myTempChar == '9'):
            myStrNum[currentInd] = '0'
            rollover = True
        elif(not myTempChar.isalnum()):
            #skip this special character
            rollover = True
        else:
            myStrNum[currentInd] = chr(ord(str(myStrNum[currentInd])) + 1)
        currentInd = currentInd - 1
    #Turn the list of characters back into a string and return it...
    return "".join(myStrNum)

def incrGenFileName(myOldFilename):
    #using ZWA as a unique ending to prevent sequential file overwrites.
    return(incrStrNum(myOldFilename[:-4], ()) + "ZWA.txt")

#In general, the UAN analysis and generation might be somewhat slow, so here are some possibilities for improvement:
#passing previously-found set lists and so on to TLP_Enrollments_Electric that might try to regenerate work that's already been done(in other words, reduce redundancy)
#precompute common sets / set operations - like for [0-9]
#see if there are more efficient ways of addressing / generating / searching etc. this set and list information.
#use lists instead of sets, maybe sortedLists with low find times.  I don't know how well that would play with Python's pickle function if that's used later, though.

#Note: this doesn't act exactly like incrStrNum.  If there's a set that contains A-Z but also a-z and 0-9, uppercase Z could become a lowercase a or a 0.  It doesn't stay in a human-imagined subset, it just
#goes to the next in the set the computer has- or loops back to the beginning of the set if it's at the end.
def findNextInSet(mySet, myChar):
    #print("\nfindingNextInSet\n")
    #return a list of the character and the rollover state
    #tempList = list(mySet)
    tempList = sorted(mySet)
    #tempStr = ''.join(mySet)
    tempStr = ''.join(tempList)
    myPos = tempStr.find(myChar)
    if((myPos + 1) < len(tempStr)):
        return [tempStr[myPos+1], False]
    else:
        return [tempStr[0], True]

#Creating this is left as a later exercise.  For now, the database seems to have account lengths that do reflect the mask length.
#def checkMaskLengthIssuesInUtilities():

#create a set of characters from a string of either a single character, or two characters with a dash between them.
#The dash means to put those characters, and every character inbetween, in the set- and leave the dash out of the set.
def processSetString(myString):
    #print("\nprocessingSetString\n")
    mySet = set()
    if(myString is None):
        sys.exit("\nInvalid setString passed.\n")
    if(len(myString) < 1):
        return mySet
    myList = myString.split('-')
    if(len(myList) == 1):
        return {myList[0]}
    elif(len(myList) > 2):
        sys.exit("\nInvalid setString passed- there should only be one dash in a comma sequence.\n")
    else:
        myStart = ord(myList[0])
        myEnd = ord(myList[1]) + 1
        for i in range(myStart, myEnd):
            mySet.add(chr(i))
    return mySet

#What about more than one dash? etc.  Check for weird stuff.
def getValidUANMaskCharacterSets(utilityCode):
    #print("\nGettingValidUANMaskCharacterSets\n")
    #myQueryString = "select AccountLength, AccountMask, AccountMaskSQL, * from DataServices%s.dbo.Utilities where UtilityCode = '%s'" % (GenericSettings.getMyEnvironment(), utilityCode)
    result = utilityMaskSearch(utilityCode)
    #else:
    myResult = result[0][2]
    #if(len(myResult) < 2):
    #list of valid sets for each character
    validCharacterResultSets = []
    #Begin string parsing with state machine
    #states: character, comma, dash, left bracket, right bracket
    if((myResult == None) or (len(myResult) < 1)):
        return []
    myIndex = 0
    #machineState = ' '
    bracketsEngaged = False
    #incremented until you get to a comma, then processed into tempSet and reset.
    setString = ''
    #tempSet accumulates all the comma-connected subsets for a single character.  After a bracket is found, tempSet is appended to validCharacterResultSets.  At that point, tempSet is reset.
    tempSet = set()
    if((myResult[myIndex] == ',') or (myResult[myIndex] == '-') or (myResult[myIndex] == ']')):
        sys.exit("\nInvalid start to the Mask String for the UAN.\n")
        #return False
    if(myResult[myIndex] != '['):
        #print("\nAppended in 3\n")
        validCharacterResultSets.append({myResult[myIndex]})
        #setString = setString + myResult[myIndex]
    else:
        machineState = myResult[myIndex]
        bracketsEngaged = True
    myIndex = myIndex + 1
    #states: character, comma, dash, left bracket, right bracket
    while(myIndex < len(myResult)):
        if myResult[myIndex] == ',':
            if(not bracketsEngaged):
                sys.exit("\nMask String commas should be between brackets.\n")
            tempSet.update(processSetString(setString))
            setString = ''
        elif myResult[myIndex] == '-':
            if(not bracketsEngaged):
                sys.exit("\nMask String dashes should be between brackets.\n")
            setString = setString + myResult[myIndex]
        elif myResult[myIndex] == '[':
            if(bracketsEngaged):
                sys.exit("\nMask String left brackets should not be between brackets.\n")
            bracketsEngaged = True
        elif myResult[myIndex] == ']':
            if(not bracketsEngaged):
                sys.exit("\nMask String right brackets should not be encountered before left brackets.\n")
            #print("\nThis is processSetString: " + str(processSetString(setString)) + "\n")
            tempSet.update(processSetString(setString))
            #print("\nThis is tempSet: " + str(tempSet) + "\n")
            #print("\nAppended in 2\n")
            validCharacterResultSets.append(tempSet)
            tempSet = set()
            setString = ''
            bracketsEngaged = False
        else:
            if(not bracketsEngaged):
                #print("\nAppended in 1\n")
                validCharacterResultSets.append({myResult[myIndex]})
            else:
                setString = setString + myResult[myIndex]
        #machineState = myResult[myIndex]
        myIndex = myIndex + 1
    if(bracketsEngaged):
        sys.exit("\nAt the end of Mask String, missing a right aka closing bracket.\n")
    return validCharacterResultSets

#Need to cover the '&' case- which stands for [a-z,A-Z,0-9]
#Also need to treat all characters to the left of the first nonzero character- from right to left- as literal.  Zeroes on that side are literal.
#Maybe also implement a buffer to guard against literal zeroes that aren't marked by nonzero characters around them.  But if you interpret that wrong, you could interpret correct existing uan's as incorrect.
#Interesting test cases- utility codes 69(3 nulls), 1(no mask sql but has a second mask string), 8(has acct length but no mask strings), 2(has all 3), 49(big mask sets) and 5 is the literal vs. nonliteral zeroes case
#for the second mask.
#49 also has the '&' second mask case.  And 5 has the account mask(2nd string) with literal AND nonliteral zeroes.
#Need to test for ghostliterals(ghost literal zeroes)... uan's that don't really allow nonzero characters, especially at the beginning of the string.  Utility code 1 is the only one where
#you could test for ghost literals(ghost literal zeroes)
def getAltValidUANMaskCharacterSets(utilityCode):
    #print("\nGettingAltValidUANMaskCharacterSets\n")
    #myQueryString = "select AccountLength, AccountMask, AccountMaskSQL, * from DataServices%s.dbo.Utilities where UtilityCode = '%s'" % (GenericSettings.getMyEnvironment(), utilityCode)
    result = utilityMaskSearch(utilityCode)
    #else:
    myResult = result[0][1]
    validCharacterResultSets = []
    if((myResult == None) or (len(myResult) < 1)):
        return []
    myIndex = len(myResult) - 1
    literalsOnly = False
    while (myIndex >= 0):
        if(myResult[myIndex] == '0'):
            if(not literalsOnly):
                validCharacterResultSets.append({'0', '1', '2', '3', '4', '5', '6', '7', '8', '9'})
            else:
                validCharacterResultSets.append({myResult[myIndex]})
        elif(myResult[myIndex] == '&'):
            if(not literalsOnly):
                myTempSet = processSetString("a-z")
                myTempSet.update(processSetString("A-Z"))
                myTempSet.update(processSetString("0-9"))
                validCharacterResultSets.append(myTempSet)
            else:
                validCharacterResultSets.append({myResult[myIndex]})
        else:
            validCharacterResultSets.append({myResult[myIndex]})
            literalsOnly = True
        myIndex = myIndex - 1
    return validCharacterResultSets

#only checks for the validity of the uan against the utility code's rules arounds masks and account length, etc.
def validateUAN(UAN, utilityCode):
    #print("\nvalidatingUAN\n")
    #repeating the query potentially three times... might just want to return it in some kind of set list or something.  Don't know how common repeated queries would be,
    #but it seems like a worthwhile optimization.
    if ((UAN is None) or (len(UAN) < 1)):
        return False
    myListOfSets = getValidUANMaskCharacterSets(utilityCode)
    if((myListOfSets is None) or (len(myListOfSets) < 1)):
        # If accountMask exists use that- whereever a character isn't a zero, use it as-is, but where it is a zero, generate a [0-9] character.
        # though some of them may actually be zeroes.  But this should be safe, I think.
        myAltListOfSets = getAltValidUANMaskCharacterSets(utilityCode)
        if ((myAltListOfSets is None) or (len(myAltListOfSets) < 1)):
            # elif account length exists use that(though there isn't currently an account-length-only entry)
            #myQueryString = "select AccountLength, AccountMask, AccountMaskSQL, * from DataServices%s.dbo.Utilities where UtilityCode = '%s'" % (GenericSettings.getMyEnvironment(), utilityCode)
            result = utilityMaskSearch(utilityCode)
            myResult = result[0][0]
            if ((myResult == None) or (len(myResult) < 1)):
                #situation is: no mask sets, no mask, no account length.  So as long as the UAN isn't none and has a length > 0, which is established above, it's a valid UAN.
                return True
            else:
                #Considering account length.
                if(len(UAN) == int(myResult)):
                    return True
                else:
                    return False
        else:
            if(len(UAN) != len(myAltListOfSets)):
                return False
            anIndex = 0
            while anIndex < len(myAltListOfSets):
                if(UAN[anIndex] not in myAltListOfSets[anIndex]):
                    return False
                anIndex = anIndex + 1
            return True
    else:
        if (len(UAN) != len(myListOfSets)):
            return False
        for i in range(0,len(myListOfSets)):
            if(UAN[i] not in myListOfSets[i]):
                return False
        return True

#Does all of the below while checking to make sure it's compatible with the utility's UAN rules.
#takes an alphanumeric string with numbers and letters and increments it... if the last letter is a num, it ups that 'til
#it rolls over into the next num over.  Same with a letter, but in the letter space... and if there's rollover, it checks the next
#slot over and makes sure that one stays the same type that it was originally.
#This function assumes that if it is not being sent None or an empty string, then the UAN should be otherwise valid.
#only checks to ensure continued UAN validity, not uniqueness in the program / lack of collisions against the database.
def safelyIncrValidUAN(someStrNum, utilityCode, uniqueUANSet):
    #print("\nsafelyIncrementingUAN\n")
    if((someStrNum is None) or (len(someStrNum) < 1)):
        #Garbage in, garbage out.  We can't say for sure the person running this function wants the program to crash on a null / NONE / empty string, so instead we'll just pass it right back.
        #return someStrNum
        # UAN is of an inappropriate length, therefore generate a valid UAN.  It doesn't have to be unique here, there are checks for that elsewhere.
        # This should never happen as the function is not meant to handle invalid uan's, as stated above the function.
        return generatePerfectUAN(utilityCode, uniqueUANSet)
        #sys.exit("\nInvalid string sent to safelyIncrValidUAN- string is None or length 0.\n")
    #turn the string-num into a list of characters, because Python's strings are immutable.
    myStrNum = list(someStrNum)
    currentInd = len(myStrNum) - 1
    #print("\nThis is currentInd: " + str(currentInd) + "\n")
    rollover = True
    myTempChar = ''
    myTentativeChar = ''
    mySetList = getValidUANMaskCharacterSets(utilityCode)
    if (len(mySetList) < 1):
        mySetList = getAltValidUANMaskCharacterSets(utilityCode)
        if (len(mySetList) < 1):
            #myQueryString = "select AccountLength, AccountMask, AccountMaskSQL, * from DataServices%s.dbo.Utilities where UtilityCode = '%s'" % (GenericSettings.getMyEnvironment(), utilityCode)
            result = utilityMaskSearch(utilityCode)
            myResult = result[0][0]
            if(myResult == None):
                #situation is: no mask sets, no mask, no account length.  So as long as the UAN isn't none and has a length > 0, which is established above, it's a valid UAN.
                return incrStrNum(someStrNum, ())
            else:
                #Considering account length.
                if(len(someStrNum) == int(myResult)):
                    return incrStrNum(someStrNum, ())
                else:
                    #UAN is of an inappropriate length, therefore generate a valid UAN.  It doesn't have to be unique here, there are checks for that elsewhere.
                    #This should never happen as the function is not meant to handle invalid uan's, as stated above the function.
                    return generatePerfectUAN(utilityCode, uniqueUANSet)
    while(rollover and (currentInd >= 0)):
        #print("\nThis is currentInd: " + str(currentInd) + "\n")
        rollover = False
        myTempChar = str(myStrNum[currentInd])
        someList = findNextInSet(mySetList[currentInd], myTempChar)
        #myTempChar = someList[0]
        myStrNum[currentInd] = someList[0]
        rollover = someList[1]
        currentInd = currentInd - 1
    #Turn the list of characters back into a string and return it...
    return "".join(myStrNum)

#takes a UAN that follows its utilityCode's rules about mask, length etc, and then makes sure that it is unique in the running program(among the list of UAN's being processed) and in the dev / test environment database(whether qa or pt)
def makeValidUANPerfect(sentUAN, utilityCode, uniqueUANSet):
    uanImperfect = True
    #fix for possible issue with reassigning supplied function variables in a loop - memory scoping or whatnot.
    #Always increment first, in case you mock before something has posted to the database.
    tempUAN = safelyIncrValidUAN(sentUAN, utilityCode, uniqueUANSet)
    while(uanImperfect):
        #print("\nThis is tempUAN: " + str(tempUAN) + "\n")
        if tempUAN in uniqueUANSet:
            tempUAN = safelyIncrValidUAN(tempUAN, utilityCode, uniqueUANSet)
            continue
        myQueryString = "select * from Enrollment%s.dbo.InboundData where UtilityAccountNumber='%s'" % (GenericSettings.getMyEnvironment(), tempUAN)
        result = GenericSettings.genericSQLQuery(myQueryString)
        #the environment is only in the hostname, not the databases, so we don't have to deal with that in Postgres.
        if(GenericSettings.getMyEnvironment() == "PT"):
            #PT->Databases->nerf->schemas->public→Tables→nerf_orderitem
            pgQueryString = "SELECT * FROM nerf.public.nerf_orderitem WHERE uan = '%s'" % tempUAN
        elif(GenericSettings.getMyEnvironment() == "QA"):
             #QA->Databases→ebdb->Schemas->public->Tables->nerf_orderitem
            pgQueryString = "SELECT * FROM ebdb.public.nerf_orderitem WHERE uan = '%s'" % tempUAN
        else:
            sys.exit("\nEnvironment not recognized in makeValidUANPerfect.  GenericSettings.getMyEnvironment() is: " + str(GenericSettings.getMyEnvironment()) + "\n")
        pgResult = GenericSettings.genericPGSQLQuery(pgQueryString)
        if ((len(result) <= 0) and (len(pgResult) <= 0)):
            uanImperfect = False
            #print("\nReturning the tempUAN in makeValidUANPerfect\n")
            return tempUAN
        else:
            tempUAN = safelyIncrValidUAN(tempUAN, utilityCode, uniqueUANSet)

#This function isn't concerned with whether the UAN already exists- that's handled in another function.
def generatePerfectUAN(utilityCode, uniqueUANSet):
    #print("\ngeneratingValidUAN\n")
    #repeating the query potentially three times... might just want to return it in some kind of set list or something.  Don't know how common repeated queries would be,
    #but it seems like a worthwhile optimization.
    tempUAN = ''
    while(True):
        myListOfSets = getValidUANMaskCharacterSets(utilityCode)
        if((myListOfSets is None) or (len(myListOfSets) < 1)):
            # If accountMask exists use that- whereever a character isn't a zero, use it as-is, but where it is a zero, generate a [0-9] character.
            # though some of them may actually be zeroes.  But this should be safe, I think.
            myAltListOfSets = getAltValidUANMaskCharacterSets(utilityCode)
            if ((myAltListOfSets is None) or (len(myAltListOfSets) < 1)):
                # elif account length exists use that(though there isn't currently an account-length-only entry)
                #myQueryString = "select AccountLength, AccountMask, AccountMaskSQL, * from DataServices%s.dbo.Utilities where UtilityCode = '%s'" % (GenericSettings.getMyEnvironment(), utilityCode)
                result = utilityMaskSearch(utilityCode)
                myResult = result[0][0]
                #if account length nonexistent:
                if ((myResult == None) or (len(myResult) < 1)):
                    #situation is: no mask sets, no mask, no account length.  So as long as the UAN isn't none and has a length > 0, which is established above, it's a valid UAN.
                    #As this is such a special case, generating a unique uan could be a pain, at least in terms of processing time.
                    #if tempUAN not in uniqueUANSet:
                    #    return tempUAN
                    #else:
                    tempUAN = '000000001'
                else:
                    #account length section
                    for a in range(0, int(myResult)):
                        tempUAN = tempUAN + '0'
                    #return tempUAN

            else:
                for i in range(0, len(myAltListOfSets)):
                    # pick a random member of the set and add that character to the string
                    tempUAN = tempUAN + random.choice(tuple(myAltListOfSets[i]))
                #return tempUAN
        else:
            for i in range(0, len(myListOfSets)):
                # pick a random member of the set and add that character to the string
                #print("\nThis is myListOfSets: " + str(myListOfSets) + "\n")
                #print("\nThis is myListOfSets[i]: " + str(myListOfSets[i]) + "\n")
                tempUAN = tempUAN + random.choice(tuple(myListOfSets[i]))
            #return tempUAN
        #if tempUAN not in uniqueUANSet:
        #    return tempUAN
        #else:
        #    tempUAN = tempUAN
        tempUAN = makeValidUANPerfect(tempUAN, utilityCode, uniqueUANSet)
        return tempUAN

def mockUANUntilPerfect(myContainerList, myUtilityCode, uniqueUANSet):
    if(not validateUAN(myContainerList[0], myUtilityCode)):
        myContainerList[0] = generatePerfectUAN(myUtilityCode, uniqueUANSet)
    else:
        myContainerList[0] = makeValidUANPerfect(myContainerList[0], myUtilityCode, uniqueUANSet)
    uniqueUANSet.add(myContainerList[0])
    if(len(myContainerList) == 3):
        myContainerList[2] = uniqueUANSet
    else:
        myContainerList.append(uniqueUANSet)
    return myContainerList

def mockUIDUntilPerfect(myContainerList, uniqueUIDSet):
    myServer = GenericSettings.liveServer

    UIDLoop = True
    while (UIDLoop):
        #print("\nCheck for an available UID.\n")
        UIDLoop = False
        #someOrderedDict['UID'][someListIndex] = incrStrNum(someOrderedDict['UID'][someListIndex], ())
        myContainerList[1] = incrStrNum(myContainerList[1], ())
        if(myContainerList[1] in uniqueUIDSet):
            UIDLoop = True
            continue
        myQueryString = "select * from Enrollment%s.dbo.InboundData where UID='%s'" % (GenericSettings.getMyEnvironment(), myContainerList[1])
        result = GenericSettings.genericSQLQuery(myQueryString)
        if len(result) > 0:
            UIDLoop = True
            continue
        GenericSettings.safelySetMSSQLConnection('WNTEPNSQLD1')
        myQueryString = "select * from EnrollmentDEV.dbo.InboundData where UID='%s'" % myContainerList[1]
        result = GenericSettings.genericSQLQuery(myQueryString)
        if len(result) > 0:
            UIDLoop = True
        GenericSettings.safelySetMSSQLConnection(myServer)
    uniqueUIDSet.add(myContainerList[1])
    myContainerList[3] = uniqueUIDSet
    return myContainerList

#It makes the most sense to do the perfection check for each uan the second they come in, and then use a list of "perfect" uan's for the rest of the process.

#myContainerList should have two elements, UAN then UID, and the list is used to pass-by-reference in Python.
#Probably a simple matter of writing safelyIncrUAN, then checking for UAN validity at the beginning of the function and changing incrStrNums to safelyIncrUAN's.

#Perfect here means it is a valid uan per the utility(integrity), it doesn't preexist in the relevant server's database(no collisions), and it isn't one of the uan's already used.
#It's worth noting that the UID's may not have to be unique at all.  I'm pretty sure I've seen textfile with lots of copies.  I'll have to confirm that, though.
#"Perfect" UANs still have to be inserted very carefully, because if they're re-checked, of course they will no longer be unique- they will have been put in there earlier!
#Also the whole idea with maintaining a list and a set of uan's was so that when set insertion failed, that uan had to be redone... but what is the most efficient and least confusing way to do this?  When do we do the uniqueness check, and the
#collisions check and the validity check?
#And oh yes, the uniqueUANSet was supposed to be in container[2] now and if I have a unique uid set that would go in [3].
#And if I handed a blank uniqueUANSet to this function, that was so the programs with less complicated constraints could use it.  But I'm moving most of them towards being that complicated, with file lists and such, I think.
def mockUANAndUIDUntilPerfect(myContainerList, myUtilityCode, uniqueUANSet, uniqueUIDSet):
    #insertion of the new uan into the unique uan set is accomplished in the below function.
    if(len(myContainerList) == 3):
        myContainerList.append(uniqueUIDSet)
    elif (len(myContainerList) == 2):
        myContainerList.append(uniqueUANSet)
        myContainerList.append(uniqueUIDSet)
    myContainerList = mockUANUntilPerfect(myContainerList, myUtilityCode, uniqueUANSet)
    myContainerList = mockUIDUntilPerfect(myContainerList, uniqueUIDSet)
    return myContainerList

def getUtilityCodeFromUtilitySlug(myUtilitySlug):
    myQueryString = "select UtilityCode from DataServices%s.dbo.Utilities where UtilitySlug='%s'" % (GenericSettings.getMyEnvironment(), myUtilitySlug)
    result = GenericSettings.genericSQLQuery(myQueryString)
    if(len(result) > 0 and len(result) < 2):
        return result[0][0]
    else:
        sys.exit("\nCouldn't find a utility code for the SKU in question.  This case should never occur.\n")

def blankOrNoneCheck(myStr):
    if ((myStr is None) or (myStr == '')):
        return True
    elif (myStr.strip() == ''):
        return True
    else:
        return False

#Tries to assign the first key's cell's contents to the second key's cell's contents first if they're not equal and the first key cell isn't a blank string or None.  Tries assigning the second key cell to the first if that's not the case and they're not equal. 
def makeEqual(myDict, myKey, myKey2, myRow):
    if (myDict[myKey][myRow] != myDict[myKey2][myRow]):
        if (not blankOrNoneCheck(myDict[myKey][myRow])):
            myDict[myKey2][myRow] = myDict[myKey][myRow]
        else:
            myDict[myKey][myRow] = myDict[myKey2][myRow]
    return myDict

#localSystemBasePath can be gleaned from:
#    GenericSettings.initializeConn()
#    myBasePath = GenericSettings.getTheBasePath()
#    For example, localSystemBasePath for me is:
#    C:\Users\mhissong\Desktop\Regression\Regression\ECC\TestUtilities\
def standardizeAnExcelOrTLPFile(localSystemBasePath, fullFilePath):
    myFilePath = os.path.split(fullFilePath)[0] + os.sep
    myFileName = os.path.split(fullFilePath)[-1]
    altFullFilePath = myFilePath + "B" + myFileName
    myString = localSystemBasePath + "UtilitySupportModules\\ExcelFileStandardizer.vbs \"" + myFilePath + "\" \"" + myFileName 
    os.system(myString)
    os.remove(fullFilePath)
    os.rename(altFullFilePath, fullFilePath)

#Slightly modified from https://stackoverflow.com/questions/23861680/convert-spreadsheet-number-to-column-letter
def arrayIndexToExcelColumnLetters(n):
    n = n + 1
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string

#From: https://stackoverflow.com/questions/50917384/read-drop-down-values-from-excel-using-pandas
#Shows ways to handle dropdowns in Excel spreadsheets.
#run the VB, capture its stdout output and return it as a Python list.  It's possible that whitespace could get garbled with this method.
#I'll try this for a first iteration.
#docs.python.org/2/library/commands.html – Eric Leschinski Jul 14 '16 at 15:05 
def readListFromDropdown(book_name, sheet_name, range_str):
    pass
    
    #This VB should work.  Should be called with os.system(myCmd).  Or this whole program should be written in VB, which I'm leaning towards at present.
    #Sub Iterate_Through_data_Validation()
    #    Dim xRg As Range    
    #    Dim xCell As Range    
    #    Dim xRgVList As Range    
    #    Set xRg = Worksheets("Sheet1").Range("B8")    
    #    Set xRgVList = Evaluate(xRg.Validation.Formula1)    
    #    For Each xCell In xRgVList    
    #        xRg = xCell.Value    
    #        ActiveSheet.PrintOut    
    #    Next    
    #End Sub
    
    #os.system("ls") is for running a command where you only care if it ran or not. If you want the stdout output,
    #use something like commands.getstatusoutput("ls")[1] as defined here:
    #docs.python.org/2/library/commands.html – Eric Leschinski Jul 14 '16 at 15:05 

    #Doesn't work.  The drop down is supposedly proprietary.  VB, C# are supposed to work.  Maybe Java through OpenOffice.
    #wb = openpyxl.load_workbook(book_name)
    #ws = wb[sheet_name]
    #data = [[cell.value for cell in row] for row in ws[range_str]]
    #
    #validations = ws.data_validations.dataValidation
    #for validation in validations:
    #    ranges = validation.sqref.ranges
    #    if len(ranges) != 1:
    #        raise NotImplementedError
    #    if validation.type == 'list':
    #        print(str(validation))
    #        print(str(validation.formula1))
    #        list_cells = list(ws[validation.formula1])
    #        values = [cell.value for cell_row in list_cells for cell in cell_row]
    #    else:
    #        raise NotImplementedError
    #    bounds = ranges[0].bounds
    #    try:
    #        data[bounds[1]-1][bounds[0]-1] = values
    #    except IndexError:
    #        pass
    #return data

    #The VB Alternative:
    # This VB should work.  Should be called with os.system(myCmd).  Or this whole program should be written in VB, which I'm leaning towards at present.
    # Sub Iterate_Through_data_Validation()
    #    Dim xRg As Range    
    #    Dim xCell As Range    
    #    Dim xRgVList As Range    
    #    Set xRg = Worksheets("Sheet1").Range("B8")    
    #    Set xRgVList = Evaluate(xRg.Validation.Formula1)    
    #    For Each xCell In xRgVList    
    #        xRg = xCell.Value    
    #        ActiveSheet.PrintOut    
    #    Next    
    # End Sub
    
    #For Reading Output, if working with a vb os.system-ish program call:
    # os.system("ls") is for running a command where you only care if it ran or not. If you want the stdout output,
    # use something like commands.getstatusoutput("ls")[1] as defined here:
    # docs.python.org/2/library/commands.html – Eric Leschinski Jul 14 '16 at 15:05 

#not guaranteed to be interchangeable with other ways of creating a TLP.  Needs testing.
def alternate_read_a_TLP_File(myFile):
    #Might have to use openpyxl for handling the product builder.  To do that, reference "AddProductFromError.py".

    currFullFilename = myFile
    #currFullFilename = GenericSettings.getTheBasePath() + "TLPFileCreator" + os.sep + "TLPFileResult0.txt"
    #currFullFilename = "C:os.sepUsersos.sepmhissongos.sepDesktopos.sepRegressionos.sepRegressionos.sepECCos.sepTestUtilitiesos.sepTLPFileCreatoros.sepTLPFileResult0.txt"
    #open(filename, encoding="utf8")
    
    #Note: this encoding is for the SAP product regression builder.  Other TLP's are probably stored in the default "cp1252" format.  I don't know why the product builder TLP's are stored this way.  Maybe it has something to do with...
    #Wait.  It's the linux(Mac) thing... when they're uploaded by Macs or linux.  The file is somehow compressed or stored differently.  I'll probably still run into the rest of the algorithm issues, but I should open and resave each of these first.
    #Like I did for the TLPProductErrorCorrection stuff.
    #myFilePointer = open(currFullFilename, "r", encoding="latin-1")
    myFilePointer = open(currFullFilename, "r")
    
    #myFileString = myFilePointer.read().strip()
    myList = myFilePointer.readlines()
    myFilePointer.close()
    #myList = myFileString.split("\n")
    rows = []
    myOrderedDict = OrderedDict()
    #x = 0
    for i in myList:
        #print("/nThis is myList[" + str(x) + "]: " + myList[x] + "\n")
        tempList = i.strip("\n").split("\t")
        rows.append(tempList)
        #x = x + 1
    #goldenRow = rows[-1]
    #del rows[-1]
    colLen = len(rows)
    rowLen = len(rows[0])
    for y in range(0, rowLen):
        tempList = []
        for x in range(1,colLen):
            #print("\nThis is x,y: " + str(x) + "," + str(y) + "\n")
            tempList.append(rows[x][y])
        myOrderedDict.update({rows[0][y]: tempList})
    myReturnList = [myOrderedDict, rowLen, colLen]
    return myReturnList
    #myReturnList.append(myOrderedDict)
    #return myOrderedDict

#not guaranteed to be interchangeable with other ways of writing a TLP.  Needs testing.
def alternate_write_a_TLP_File(myOrderedDictElementsList, currOutputFilename):
    myOrderedDict = myOrderedDictElementsList[0]
    myKeys = list(myOrderedDict.keys())
    colLen = myOrderedDictElementsList[2]
    myFilePointer = open(currOutputFilename, "w")
    for i in myKeys:
        myFilePointer.write(i + "\t")
    myFilePointer.write("\n")
    for x in range(0, (colLen-1)):
        for y in myKeys:
            myFilePointer.write(str(myOrderedDict[y][x]) + "\t")
        myFilePointer.write("\n")
    myFilePointer.close()
    #pass
    #myFilePointer = open(currOutputFilename, "w")
    #for i in myKeys:
    #    myFilePointer.write(i + "\t")
    #myFilePointer.write("\n")
    #for x in range(0, (colLen-1)):
    #    for y in myKeys:
    #        myFilePointer.write(str(myOrderedDict[y][x]) + "\t")
    #    myFilePointer.write("\n")
    #myFilePointer.close()
    ##Don't necessarily know the dictionary will have even a single 'value' row.
    ##myReturnList = [myNewRelativeFilename, myOrderedDict['UtilityAccountNumber'][0], uniqueUANSet, uniqueUIDSet]
    #backupTLPFilePath = backupTLPDir + relativeFilename
    ##print("\nThis is backupTLPFilePath: " + backupTLPFilePath + "\n")
    #shutil.move(currFullFilename, backupTLPFilePath)

#might want to create a Python TLP Ordered Dict class so that I inherently have colLen, rowLen and myKeys.
#All this because pandas and whatever reader I was using weren't playing nicely with my TLP's.  Geez.
#maybe with a "find all" flag instead of just looking for the first instance.
def find_In_An_Ordered_Dict(colLen, myKeys, myOrderedDict, searchObject):
    for x in range(0, (colLen-1)):
        for y in myKeys:
            if(str(myOrderedDict[y][x]) == searchObject):
                return [y,x]
