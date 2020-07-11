import GenericSettings
import TLPSupportModule
import ProductBuilderStatusPage
import os
import sys
import shutil
from shutil import copyfile
import openpyxl
from datetime import datetime
#From https://www.seleniumeasy.com/python/example-code-using-selenium-webdriver-python
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.remote import webelement
import time
import pyautogui
import atexit
import password_manager

global myDriver

#Adding a uniqueCounter to prevent operations occurring in the same second from resulting in the same identifier
global uniqueCounter
global usedSKUs

def exit_handler():
    try: globals()['myDriver'].close()
    except: pass

def testSetup():
    globals()['myDriver'] = webdriver.Chrome(r'C:\Users\drivers\chromedriver.exe')
    globals()['myDriver'].maximize_window()
    atexit.register(exit_handler)

def channelIsAllowed(workbook, supposedChannel):
    if(supposedChannel is None):
        supposedChannel = ""
    supposedChannel = supposedChannel.strip()
    sheet = workbook['Lookup']
    myColumnNum = 3
    if(sheet.cell(row=1, column=myColumnNum).value != "Channel Slugs"):
        sys.exit("\nFix the Channel Slugs column position finder in ProductBuilderRegressionAutomation_SAPVersion.py.\n")
    count = 2
    notDone = True
    while(notDone):
        loopVal = sheet.cell(row=count, column=myColumnNum).value
        if(loopVal is None):
            return False
        loopVal = loopVal.strip()
        if(loopVal == supposedChannel):
            return True
        if(loopVal == ""):
            return False
        count = count + 1

#validation error(s):
def checkWebPageForTextAndMaybeClick(driver, text, expectTextBool, fatalBool, delayInSeconds, waitIfUnexpectedBool, clickBool):
    time.sleep(1)
    time.sleep(delayInSeconds)
    resultBool = True
    try:
        #from: https://www.quora.com/How-do-I-find-an-element-that-contains-specific-text-in-Selenium-Webdriver-Python
        result = driver.find_element_by_xpath("//*[contains(text(), text)]")
        type = "notFound"
        storedText = ""
        storedTextSet = False
        if(isinstance(result, list)):
            if (len(result) > 0):
                type = "occupiedList"
                #print("\noccupiedList\n")
                storedText = result[0].text
                storedTextSet = True
            else:
                #print("\nemptyList\n")
                type = "emptyList"
        elif(result is None):
            type = "None"
        elif(isinstance(result, int)):
            type = "int"
        elif (isinstance(result, dict)):
            if(len(result) == 0):
                type = "emptyDict"
            else:
                type = "occupiedDict"
                #todo: retrieve text from occupiedDict.  Not critical for product builder testing, though.
        else:
            type = "webElement"
            storedText = result.text
            storedTextSet = True
        if(clickBool):
            if(type == "occupiedList"):
                result[0].click()
            elif(type == "WebElement"):
                result.click()
        ##print("\nThis is type: " + type + "\n")
        resultBool = False
        if((type != "None") and (type != "emptyList") and (type != "emptyDict")):
            if(storedTextSet):
                ##print("\nstoredText is: " + storedText + "\n")
                ##print("\ntext is: " + text + "\n")
                if(text in storedText):
                    resultBool = True
    except:
        resultBool = False
    if(resultBool == expectTextBool):
        return True
    else:
        ##print("\nSearched text is: " + text + "\n")
        startingText = "\ncheckWebPageForTextAndMaybeClick failed, expecting this error state: " + str(expectTextBool) + " when it got the oppposite.  "
        if(waitIfUnexpectedBool):
            throwaway = input(startingText + "Press any key to continue.\n")
        else:
            ##print(startingText + "\n")
            pass
        if(fatalBool):
            sys.exit("\ncheckWebPageForTextAndMaybeClick failed, expecting this error state: " + str(expectTextBool) + " when it got the oppposite.\n")
        else:
            return False

#Allows for 3 download attempts, each lasting maxDownloadTimeInSeconds
def tryToDownload(seleniumDownloadButton, downloadDir, fileEnding, maxDownloadTimeInSeconds, fatalBool, fatalMessage):
    for i in [1,2,3]:
        currentTimeSinceEpoch = time.time()
        ##print("\nThis is currentTimeSinceEpoch: " + str(currentTimeSinceEpoch) + "\n")
        seleniumDownloadButton.click()
        fileNotFoundYet = True
        timeSpent = 0
        while(timeSpent < maxDownloadTimeInSeconds and fileNotFoundYet):
            myFileName = GenericSettings.getNameOfNewestSurfaceFileInDir(downloadDir)
            ##print("\nThis is myFileName.split(\".\"): " + str(myFileName.split(".")) + "\n")
            if(myFileName.split(".")[-1] == fileEnding):
                if(os.path.getmtime(downloadDir + myFileName) > currentTimeSinceEpoch):
                    return myFileName
            time.sleep(1)
            timeSpent = time.time() - currentTimeSinceEpoch
    if(fatalBool):
        print("\ntryToDownload function failure message: ")
        sys.exit(fatalMessage)
    else:
        return "tryToDownloadFailed"
    
def deleteInputFolderContents():
    myEnvironment = GenericSettings.getMyEnvironment()
    myDropdownOptionList = ["Cirro Energy", "NRG_regression Home", "Green Mountain Energy"]
    for i in myDropdownOptionList:
        outputFilePath = ""
        if (i == "Cirro Energy"):
            outputFilePath = GenericSettings.getTheBasePath() + "ProductBuilderRegressionAutomation_SAPVersion" + os.sep + "ExcelFileProductListInputs" + os.sep + "Cirro"
        elif (i == "NRG_regression Home"):
            outputFilePath = GenericSettings.getTheBasePath() + "ProductBuilderRegressionAutomation_SAPVersion" + os.sep + "ExcelFileProductListInputs" + os.sep + "NRG_regression"
        elif (i == "Green Mountain Energy"):
            outputFilePath = GenericSettings.getTheBasePath() + "ProductBuilderRegressionAutomation_SAPVersion" + os.sep + "ExcelFileProductListInputs" + os.sep + "GME"
        else:
            print("\nWARNING: Unrecognized option in the SAP Product Builder download list for " + myEnvironment + ": name not in: ")
            sys.exit(str(myDropdownOptionList))
        myFileList = [name for name in os.listdir(outputFilePath)]
        for a in myFileList:
            os.remove(outputFilePath + os.sep + a)
        outputFilePath = outputFilePath + "Fail"
        myFileList = [name for name in os.listdir(outputFilePath)]
        for a in myFileList:
            os.remove(outputFilePath + os.sep + a)

# On the product api side, you should consider whether active is false, and whether status is deactivated.
# Check out this valuable resource: https://selenium-python.readthedocs.io/locating-elements.html
# 3 downloads, 3 copy operations.
def downloadSpreadsheets():
    #testSetup()
    driver = globals()['myDriver']
    myEnvironment = GenericSettings.getMyEnvironment()
    if myEnvironment == "QA":
        driver.get("http://manage.products.qa.nrgpl.us/Builder/#/download")
    else:
        driver.get("http://manage.products.pt.nrgpl.us/Builder/#/download")
    myDropdownOptionList = ["Cirro Energy", "NRG_regression Home", "Green Mountain Energy"]
    #critical sleep- confirmed.
    time.sleep(1)
    content = driver.find_element_by_xpath("/html/body/div/div/div/div[2]/div/div/div[1]/select")
    #Doesn't work: content = driver.find_element_by_class_name('searchField ng-pristine ng-valid')
    content.click()
    outputFilePath = ""
    selector = ""
    myDLSourceFile = open(GenericSettings.getTheBasePath() + "ProductBuilderRegressionAutomation_SAPVersion" + os.sep + "DLSourceFile.txt", "r")
    myDefaultDownloadFolder = myDLSourceFile.readline().strip()
    myDLSourceFile.close()
    for i in myDropdownOptionList:
        outputFilePath = ""
        if (i == "Cirro Energy"):
            outputFilePath = GenericSettings.getTheBasePath() + "ProductBuilderRegressionAutomation_SAPVersion" + os.sep + "ExcelFileProductListInputs" + os.sep + "Cirro"
            selector = "/html/body/div/div/div/div[2]/div/div/div[1]/select/option[2]"
        elif (i == "Green Mountain Energy"):
            outputFilePath = GenericSettings.getTheBasePath() + "ProductBuilderRegressionAutomation_SAPVersion" + os.sep + "ExcelFileProductListInputs" + os.sep + "GME"
            selector = "/html/body/div/div/div/div[2]/div/div/div[1]/select/option[3]"
        elif (i == "NRG_regression Home"):
            outputFilePath = GenericSettings.getTheBasePath() + "ProductBuilderRegressionAutomation_SAPVersion" + os.sep + "ExcelFileProductListInputs" + os.sep + "NRG_regression"
            selector = "/html/body/div/div/div/div[2]/div/div/div[1]/select/option[4]"
        else:
            print("\nWARNING: Unrecognized option in the SAP Product Builder download list for " + myEnvironment + ": name not in: ")
            sys.exit(str(myDropdownOptionList))
        time.sleep(2)
        option = driver.find_element_by_xpath(selector)
        option.click()
        time.sleep(1)
        buttonSelector = "//*[@id=\"btn-download\"]"
        downloadButton = driver.find_element_by_xpath(buttonSelector)
        time.sleep(1)
        #output will be a downloaded filename if the download was successful, given that fatalBool is set to True.
        myFileName = tryToDownload(downloadButton, myDefaultDownloadFolder, "xlsx", 60, True, "Download failed for " + i + "\n")
        frontFileList = myFileName.split(".")
        frontFileName = ""
        for m in range(0, (len(frontFileList) - 1)):
            frontFileName = frontFileName + frontFileList[m]
        fullFilePath = myDefaultDownloadFolder + myFileName
        myFileName = frontFileName + "_A.xlsx"
        copyfile(fullFilePath, outputFilePath + os.sep + myFileName)
        myFileName = frontFileName + "_B.xlsx"
        shutil.move(fullFilePath, outputFilePath + "Fail" + os.sep + myFileName)
    #throwAway = input("\nMake sure the branded spreadsheets are in the right place for their brand.\n")

def addRate(driver, utility, rate, rate_code, brand, zone, effective_date, sopBool):
    #addUtilityRateXPath = "/html/body/div/div[3]/div/ul/li/a"
    #print("\nThis is utility: " + utility + "\n")
    #print("\nThis is rate: " + rate + "\n")
    #print("\nThis is rate_code: " + rate_code + "\n")
    #print("\nThis is brand: " + brand + "\n")
    #print("\nThis is zone: " + zone + "\n")
    #print("\nThis is effective_date: " + effective_date + "\n")
    #print("\nThis is sopBool: " + str(sopBool) + "\n")
    
    driver.find_element_by_xpath("/html/body/div/div[3]/div/ul/li/a").click()
    #time.sleep(1)
    utilitySlugTextBoxID = "id_utility_slug"
    driver.find_element_by_id(utilitySlugTextBoxID).send_keys(utility)
    
    rateTextBoxID = "id_rate"
    driver.find_element_by_id(rateTextBoxID).send_keys(rate)
    rateCodeTextBoxID = "id_rate_code"
    driver.find_element_by_id(rateCodeTextBoxID).send_keys(rate_code)
    brandSlugTextBoxID = "id_brand_slug"
    driver.find_element_by_id(brandSlugTextBoxID).send_keys(brand)
    zoneTextBoxID = "id_zone"
    driver.find_element_by_id(zoneTextBoxID).send_keys(zone)
    effectiveDateTextBoxID = "id_effective_date"
    driver.find_element_by_id(effectiveDateTextBoxID).send_keys(effective_date)
    isSopCheckBoxID = "id_is_sop"
    if(sopBool):
        driver.find_element_by_id(isSopCheckBoxID).click()
    #saveButtonXpath = "/html/body/div[1]/div[3]/div/form/div/div/input[1]"
    driver.find_element_by_xpath("/html/body/div[1]/div[3]/div/form/div/div/input[1]").click()

#This operates off the last line in the spreadsheet.
def ensureUtilityRateExists(sheet, columnHeaderValues):
    rowTarget = sheet.max_row
    myColumnNum = columnHeaderValues.index("Rate") + 1
    rate = sheet.cell(row=rowTarget, column=myColumnNum).value
    myColumnNum = columnHeaderValues.index("BrandSlug") + 1
    brand = sheet.cell(row=rowTarget, column=myColumnNum).value
    myColumnNum = columnHeaderValues.index("UtilitySlug") + 1
    utility = sheet.cell(row=rowTarget, column=myColumnNum).value
    myColumnNum = columnHeaderValues.index("ChannelSlug") + 1
    lookUp = sheet.cell(row=rowTarget, column=myColumnNum).value
    is_sop = "False"
    sopBool = False
    if(lookUp == "pa_standard_offer"):
        sopBool = True
        is_sop = "True"
    
    driver = globals()['myDriver']
    addressString = "http://utility.api.%s.nrgpl.us/admin/api/utilityratecode/?brand_slug=%s&rate=%s&utility_slug=%s&is_sop=%s" % (GenericSettings.getMyEnvironment().lower(), brand, rate, utility, is_sop)
    shortAddressString = "http://utility.api.%s.nrgpl.us/admin/api/utilityratecode/" % GenericSettings.getMyEnvironment().lower()
    credsFile = GenericSettings.getTheBasePath() + "utilityRateAdministration%s.txt" % GenericSettings.getMyEnvironment().lower()
    driver.get(addressString)
    if(checkWebPageForTextAndMaybeClick(driver, "Username:", True, False, 3, False, False)):
        userNamePWList = password_manager.get_username_and_pw_and_setup_if_necessary(credsFile, shortAddressString)
        driver.find_element_by_id("id_username").send_keys(userNamePWList[0])
        driver.find_element_by_id("id_password").send_keys(userNamePWList[1])
        driver.find_element_by_xpath("/html/body/div/div[2]/div/form/div[3]/input").click()
        #time.sleep(1)
        driver.get(addressString)
        #time.sleep(1)

    try:
        myWebElement = driver.find_element_by_xpath("/html/body/div/div[3]/div/div/form/p")
        if(myWebElement.text == "0 Utility Rate Codes"):
            rate_code = rate
            myColumnNum = columnHeaderValues.index("UtilityZone") + 1
            zone = sheet.cell(row=rowTarget, column=myColumnNum).value
            #effective_date = "2019-10-07"
            effective_date = GenericSettings.dateDashStringBeforeToday(0)
            addRate(driver, utility, rate, rate_code, brand, zone, effective_date, sopBool)
    except:
        pass
        
def utilityRateStub():
    driver = globals()['myDriver']
    addressString = "http://utility.api.qa.nrgpl.us/admin/api/utilityratecode/?brand_slug=cirro&rate=0.2&utility_slug=aepns&is_sop=False"
    shortAddressString = "http://utility.api.qa.nrgpl.us/admin/api/utilityratecode/"
    credsFile = GenericSettings.getTheBasePath() + "utilityRateAdministrationqa.txt"
    driver.get(addressString)
    if(checkWebPageForTextAndMaybeClick(driver, "Username:", True, False, 3, False, False)):
        userNamePWList = password_manager.get_username_and_pw_and_setup_if_necessary(credsFile, shortAddressString)
        driver.find_element_by_id("id_username").send_keys(userNamePWList[0])
        driver.find_element_by_id("id_password").send_keys(userNamePWList[1])
        driver.find_element_by_xpath("/html/body/div/div[2]/div/form/div[3]/input").click()
        #time.sleep(1)
        driver.get(addressString)
    #if(checkWebPageForTextAndMaybeClick(driver, "0 Utility Rate Codes", True, False, 3, False, False)):
    #    #print("\nSuccess!!!!!!!!!!!\n")
    myWebElement = driver.find_element_by_xpath("/html/body/div/div[3]/div/div/form/p")
    if(myWebElement.text == "0 Utility Rate Codes"):
        pass
        #print("\nSuccess!!!!!!!!!!!\n")
    #0 rates xpath:
    #/html/body/div/div[3]/div/div/form/p  

def copyRowToTheEndOfTheSpreadsheet(sheet, columnHeaderValues, rowNum, option, rowTarget, flatbillBool):
    sheet2 = None
    excelFile = None
    if(flatbillBool):
        excelFile = openpyxl.load_workbook(GenericSettings.getTheBasePath() + "ProductBuilderRegressionAutomation_SAPVersion" + os.sep + "CriticalFlatbillSpreadsheet" + os.sep + "flatbill.xlsx")
        sheet2 = excelFile['Sheet1']
    else:
        sheet2 = sheet
    if(rowTarget == 0):
        rowTarget = sheet.max_row + 1
    sheet.insert_rows(rowTarget)
    columnNum = 1
    for cell in sheet2[rowNum]:
        sheet.cell(row=rowTarget, column=columnNum).value = cell.value
        columnNum += 1
    dateTimeString = GenericSettings.getCurrentDateTimeString()
    myColumnNum = columnHeaderValues.index("Action") + 1
    if(option == "Add"):
        sheet.cell(row=rowTarget, column=myColumnNum).value = "Add"
        myColumnNum = columnHeaderValues.index("SKU") + 1
        sheet.cell(row=rowTarget, column=myColumnNum).value = ""
        myColumnNum = columnHeaderValues.index("ProductSlug") + 1
        sheet.cell(row=rowTarget, column=myColumnNum).value = GenericSettings.nStr(sheet2.cell(row=rowNum, column=myColumnNum).value) + dateTimeString + str(globals()['uniqueCounter'])
        globals()['uniqueCounter'] += 1
    elif(option == "Change"):
        #I don't think you have to deactivate an existing row to add a Change row.
        sheet.cell(row=rowTarget, column=myColumnNum).value = "Change"
    myColumnNum = columnHeaderValues.index("ProductName") + 1
    sheet.cell(row=rowTarget, column=myColumnNum).value = GenericSettings.nStr(sheet2.cell(row=rowNum, column=myColumnNum).value) + dateTimeString + str(globals()['uniqueCounter'])
    globals()['uniqueCounter'] += 1
    myColumnNum = columnHeaderValues.index("RateCategory") + 1
    sheet.cell(row=rowTarget, column=myColumnNum).value = ""
    ensureUtilityRateExists(sheet, columnHeaderValues)
    if(flatbillBool):
        excelFile.close()

def makeSureNotInFlight(sku):
    myTempString = "http://products.%s.nrgpl.us/api/v1/products/%s" % (GenericSettings.getMyEnvironment(), sku)
    myJ = GenericSettings.requestResponseChecker(myTempString, "theProductAPI")
    if("pending" in myJ['status']):
        return False
    else:
        return True

#make sure rows expected to follow the product check rules do follow them.
#This does not check for rate or promo code compliance, however.
def checkRowCompliance(rowNum, columnHeaderValues, sheet, complianceBool):
    #This is the override.  If it says we don't care about compliance(complianceBool = False), we just return true.
    if(not complianceBool):
        return True
    myColumnNum = columnHeaderValues.index("TermsOfServiceType") + 1
    if(GenericSettings.nStr(sheet.cell(row=rowNum, column=myColumnNum).value).lower() == "fixed"):
        myColumnNum = columnHeaderValues.index("EarlyCancellationFee") + 1
        ecfVal = GenericSettings.nStr(sheet.cell(row=rowNum, column=myColumnNum).value).lower()
        if(ecfVal == ''):
            return False
        elif(float(ecfVal) == 0):
            return False
    myColumnNum = columnHeaderValues.index("BrandSlug") + 1
    brand = GenericSettings.nStr(sheet.cell(row=rowNum, column=myColumnNum).value).lower().strip()
    if((brand == "nrg_residential") or (brand == "cirro")):
        myColumnNum = columnHeaderValues.index("ChannelSlug") + 1
        channel = GenericSettings.nStr(sheet.cell(row=rowNum, column=myColumnNum).value).lower().strip()
        if(channel == "retention"):
            myColumnNum = columnHeaderValues.index("ProductPath") + 1
            productPath = GenericSettings.nStr(sheet.cell(row=rowNum, column=myColumnNum).value).lower().strip()
            if(productPath == ''):
                return False
    return True

def findARowWithThisColumnValue(columnName, desiredValue, columnHeaderValues, sheet, notInFlightCheck, workbook, complianceBool):
    myColumnNum = columnHeaderValues.index(columnName) + 1
    skuColumnNum = 2
    channelColumnNum = 6
    counter = 2
    rowEnd = sheet.max_row + 1
    while(counter < rowEnd):
        myVal = sheet.cell(row=counter, column=myColumnNum).value
        if(myVal is not None):
            if(checkRowCompliance(counter, columnHeaderValues, sheet, complianceBool) ):
                if(sheet.cell(row=counter, column=myColumnNum).value.strip().upper() == desiredValue.strip().upper()):
                    channelVal = GenericSettings.nStr(sheet.cell(row=counter, column=channelColumnNum).value).strip()
                    if(channelIsAllowed(workbook, channelVal)):
                        greenlit = True
                        sku = GenericSettings.nStr(sheet.cell(row=counter, column=skuColumnNum).value).strip().lower()
                        if(notInFlightCheck):
                            if(sku == ""):
                                greenlit = False
                            else:
                                greenlit = makeSureNotInFlight(sku)
                        if(len(channelVal) < 3):
                            channelVal = "XXX"
                        if((sku not in globals()['usedSKUs']) and (sku != "") and (channelVal[:3] != "RO_") and greenlit):
                            globals()['usedSKUs'].append(sku)
                            return counter
        counter += 1
    return -1

def makeASafeChange(sheet, rowTarget, columnHeaderValues, workbook):
    rowNum = findARowWithThisColumnValue("TermsOfServiceType", "fixed", columnHeaderValues, sheet, True, workbook, True)
    if (rowNum != -1):
        copyRowToTheEndOfTheSpreadsheet(sheet, columnHeaderValues, rowNum, "Change", rowTarget, False)
        
def makeChangesSafe(fullExcelFileName):
    excelFile = openpyxl.load_workbook(fullExcelFileName)
    sheet = excelFile['NRP Product Catalog']
    rowTarget = sheet.max_row
    if(sheet.cell(row=rowTarget, column=1).value.lower() == "change"):
        sku = GenericSettings.nStr(sheet.cell(row=rowTarget, column=2).value).strip().lower()
        if((sku == "") or (not makeSureNotInFlight(sku))):
            columnHeaderValues = []
            for cell in sheet[1]:
                columnHeaderValues.append(cell.value)
            makeASafeChange(sheet, rowTarget, columnHeaderValues, excelFile)
            excelFile.save(fullExcelFileName)
    excelFile.close()
    return True

def makeSureTheValidateOnlyToggleIsOff(driver):
    try:
        #confirmed critical sleep.
        time.sleep(2)
        validateOnlyToggle = driver.find_element_by_xpath("//*[@id=\"validateOnly\"]")
        #driver.find_element_by_xpath("/html/body/div/div/div/div[2]/div[1]/div/div[1]/div/label")
        #driver.find_element_by_id('validateOnly')
        myResult = validateOnlyToggle.is_selected()
        
        if (myResult):
            # Like the other pieces, this HAS been tested.  The toggle works, although you may not run into a case where the
            # toggle is in the "ON" position by default, like here, necessitating it be turned off for this program.
            # If this toggle should always be "OFF", then consider setting the checkWebPageForTextAndMaybeClick function above
            ## to a fatalBool value of True at the end, just before the delayInSeconds- second to last parameter.

            validateOnlyToggle.click()
            time.sleep(2)
            myResult = validateOnlyToggle.is_selected()
            if (myResult):
                sys.exit("\nThe Upload Excel Product Catalog validate only toggle isn't turning Off from original On position.\n")
    except Exception as myException:
        throwaway = input("\nCaught this exception: " + repr(myException))
        sys.exit("Exiting on exception in makeSureTheValidateOnlyToggleIsOff.\n")

def doTheUpload(myButtonXpath, inputFilePath, driver, myWebAddress, expectSuccessBool):
    myFileList = [name for name in os.listdir(inputFilePath)]
    totalPath = ""
    for a in myFileList:
        notUploaded = True
        while (notUploaded):
            #if(expectSuccessBool):
            makeChangesSafe(inputFilePath + os.sep + a)
            driver.get(myWebAddress)
            time.sleep(1)
            #myResult = checkWebPageForTextAndMaybeClick(driver, "OFF", True, False, 2, True, False)
            makeSureTheValidateOnlyToggleIsOff(driver)
            myButton = driver.find_element_by_xpath(myButtonXpath)
            myButton.click()
            time.sleep(1)
            pyautogui.typewrite(inputFilePath + os.sep + a, interval=0)
            pyautogui.press('enter')
            #validation error(s):
            if(not checkWebPageForTextAndMaybeClick(driver, "validation error", not expectSuccessBool, False, 15, True, False)):
                print("\nPlease deselect the terminal window; then make any necessary fixes if this error is expected;\n")
                input("\nkeep any fixes in the same Excel file under the same name and save.  Then reselect the terminal window and press any key.\n")
            else:
                if(expectSuccessBool):
                    totalPath = inputFilePath + os.sep + a
                    print("\nThis is totalPath being sent to checkSpreadsheet: " + totalPath + "\n")
                    if(not ProductBuilderStatusPage.checkSpreadsheetVsAPIBase(driver, totalPath, a)):
                        input("\nPlease examine the errors and, if they are expected, fix the automation such that it doesn't trip over them.  Press any key to quit.\n")
                        sys.exit()
                notUploaded = False
                time.sleep(1)

def checkExpectedErrors(brand, driver):
    #RateCode
    #No RateCode found for ppl at 0.80.
    
    #/html/body/div/div/div/div[2]/div[2]/div[2]/div[2]/table/tbody/tr[1]/td/div/ul/li[1]
    #/html/body/div/div/div/div[2]/div[2]/div[2]/div[2]/table/tbody/tr[1]/td/div/ul/li[2]
    
    #
    #Channel-ProductPath, ProductPath required when Channel is Retention.
    #
    #/html/body/div/div/div/div[2]/div[2]/div[2]/div[2]/table/tbody/tr[1]/td/div/ul/li
    #
    #Partners_With_Constraints, Commodity: 'electric', Partner: 'cir' and Promotion: '555' not a valid combination.
    #
    #/html/body/div/div/div/div[2]/div[2]/div[2]/div[2]/table/tbody/tr[1]/td/div/ul/li
    #EarlyCancellationFee-PenaltyType, EarlyCancellationFee required when PenaltyType is supplied.
    #EarlyCancellationFee-TermsOfServiceType, Early Cancellation Fee required for Fixed Terms of Service.
    #
    #/html/body/div/div/div/div[2]/div[2]/div[2]/div[2]/table/tbody/tr[1]/td/div/ul/li[1]
    #/html/body/div/div/div/div[2]/div[2]/div[2]/div[2]/table/tbody/tr[1]/td/div/ul/li[2]
    time.sleep(1)
    boolOne = False
    myCount = 1
    endNotFound = True
    #print("\nIn checkExpectedErrors, brand is: " + brand + "\n")
    while(endNotFound):
        try:
            #rowXpath = "/html/body/div/div/div/div[2]/div[2]/div[2]/div[1]/div[2]/div/div[%s]/div[8]/div[2]/div/span" % myCount
            rowXpath = "/html/body/div/div/div/div[2]/div[2]/div[2]/div[1]/div[2]/div/div[%s]" % myCount
            rowElement = driver.find_element_by_xpath(rowXpath)
            if(rowElement is None):
                endNotFound = False
            else:
                myCount = myCount + 1
        except:
            endNotFound = False
    #NOTE: myCount will always be one more than the number of errors on the page because it's to be used in range() where the last iteration is one less than the end range number.
    if(((brand == "NRG_regression Home") or (brand == "Cirro Energy")) and (myCount != 5)) or ((brand == "Green Mountain Energy") and (myCount != 4)):
        throwaway = input("\nIncorrect error count for brand " + brand + " expected errors spreadsheet.  THIS IS A FAILURE.  Press any key to continue or Ctrl-C in the terminal to quit.\n")
    for count in range(1,myCount):
        try:
            #rowXpath = "/html/body/div/div/div/div[2]/div[2]/div[2]/div[1]/div[2]/div/div[%s]/div[8]/div[2]/div/span" % count
            rowXpath = "/html/body/div/div/div/div[2]/div[2]/div[2]/div[1]/div[2]/div/div[%s]" % count
            rowElement = driver.find_element_by_xpath(rowXpath)
            rowElement.click()
            time.sleep(1)
            boolOne = checkWebPageForTextAndMaybeClick(driver, "No RateCode found for", True, False, 1, False, False)
            if(not boolOne):
                boolOne = checkWebPageForTextAndMaybeClick(driver, "not a valid combination.", True, False, 1, False, False)
            if(not boolOne):
                boolOne = checkWebPageForTextAndMaybeClick(driver, "EarlyCancellationFee-PenaltyType, EarlyCancellationFee required when PenaltyType is supplied.", True, False, 1, False, False)
            if((brand == "NRG_regression Home") or (brand == "Cirro Energy")):
                if(not boolOne):
                    boolOne = checkWebPageForTextAndMaybeClick(driver, "Channel-ProductPath, ProductPath required when Channel is Retention.", True, False, 1, False, False)
            if(not boolOne):
                throwAway = input("\nUnexpected error when checking a fail spreadsheet(Product Builder negative test.  Examine the situation and press any key to go on.\n")
                #sys.exit()
        except Exception as e:
            throwAway = input("\nUnexpected exception " + str(e) + " when checking a fail spreadsheet Product Builder negative test.  Examine the situation and press any key to go on.\n")

def changeProductsToDeactivate(inputFilePath):
    myFileList = [name for name in os.listdir(inputFilePath)]
    for a in myFileList:
        workbook = openpyxl.load_workbook(inputFilePath + os.sep + a)
        sheet = workbook['NRP Product Catalog']
        columnHeaderValues = []
        for cell in sheet[1]:
            columnHeaderValues.append(cell.value)
        #findARowWithThisColumnValue(columnName, desiredValue, columnHeaderValues, sheet, notInFlightCheck, workbook, complianceBool)
        firstModificationRow = findARowWithThisColumnValue('Action', 'Add', columnHeaderValues, sheet, True, workbook, True)
        if(firstModificationRow == -1):
            workbook.close()
            sys.exit("\nNo compatible Add row found in changeProductsToDeactivate(...).  Shutting the program down.\n")
        myColumnNum = columnHeaderValues.index('Action') + 1
        rowEnd = sheet.max_row + 1
        for i in range(firstModificationRow, rowEnd):
            sheet.cell(row=i, column=myColumnNum).value = 'Deactivate'
        workbook.save(inputFilePath + os.sep + a)
        workbook.close()
        
def uploadSpreadsheetsAndCheckProducts():
    #testSetup()
    driver = globals()['myDriver']
    myWebAddress = ""
    myEnvironment = GenericSettings.getMyEnvironment()
    if myEnvironment == "QA":
        myWebAddress = "http://manage.products.qa.nrgpl.us/Builder/#/upload"
    else:
        myWebAddress = "http://manage.products.pt.nrgpl.us/Builder/#/upload"
    driver.get(myWebAddress)

    makeSureTheValidateOnlyToggleIsOff(driver)

    myBrands = ["Cirro Energy", "NRG_regression Home", "Green Mountain Energy"]

    inputFilePath = ""
    for i in myBrands:
        inputFilePath = ""
        if (i == "Cirro Energy"):
            inputFilePath = GenericSettings.getTheBasePath() + "ProductBuilderRegressionAutomation_SAPVersion" + os.sep + "ExcelFileProductListOutputs" + os.sep + "Cirro"
        elif (i == "NRG_regression Home"):
            inputFilePath = GenericSettings.getTheBasePath() + "ProductBuilderRegressionAutomation_SAPVersion" + os.sep + "ExcelFileProductListOutputs" + os.sep + "NRG_regression"
        elif (i == "Green Mountain Energy"):
            inputFilePath = GenericSettings.getTheBasePath() + "ProductBuilderRegressionAutomation_SAPVersion" + os.sep + "ExcelFileProductListOutputs" + os.sep + "GME"
        else:
            print("\nWARNING: Unrecognized option in the myBrands list for " + myEnvironment + ": name not in: ")
            sys.exit(str(myBrands))
        myButtonXpath = "/html/body/div/div/div/div[2]/div[1]/div/div[2]"
        time.sleep(2)
        
        #Maybe log the exact error encountered on the web page.  Consider automatically taking a screenshot, too?
        doTheUpload(myButtonXpath, inputFilePath, driver, myWebAddress, True)
        #changeProductsToDeactivate(inputFilePath)
        #doTheUpload(myButtonXpath, inputFilePath, driver, myWebAddress, True)
        inputFilePath = inputFilePath + "Fail"
        doTheUpload(myButtonXpath, inputFilePath, driver, myWebAddress, False)
        checkExpectedErrors(i, driver)

def updateCellsWithDateAndTime(rowTarget, columnHeaderValues, columnName, sheet):
    myColumnNum = columnHeaderValues.index(columnName)
    cellIndex = TLPSupportModule.arrayIndexToExcelColumnLetters(myColumnNum) + str(rowTarget)
    sheet[cellIndex] = sheet[cellIndex].value + GenericSettings.getCurrentDateTimeString()
    return sheet

def createAProductThroughCopyingAndModifyingTheFirstRow(fullExcelFileName):
    excelFile = openpyxl.load_workbook(fullExcelFileName)
    sheet = excelFile['NRP Product Catalog']
    rowTarget = sheet.max_row + 1
    sheet.insert_rows(rowTarget)
    
    #Here's the web address for copying a row.
    #https://pastebin.com/2LVSXTMx
    # Copy rows from original sheet and set the counter
    row_data = sheet.rows
    counter = 1
    # Version keeps track of which original row is active
    for version, row in enumerate(row_data, 1):
        if(counter == 2):
            for number, item in enumerate(row, 1):
                sheet.cell(row=rowTarget, column=number).value = item.value
            break
        #print("\nThis is counter: " + str(counter) + "\n")
        counter += 1
        
    columnHeaderValues = []
    for cell in sheet[1]:
        columnHeaderValues.append(cell.value)
    sheet = updateCellsWithDateAndTime(rowTarget, columnHeaderValues, "ProductName", sheet)
    sheet = updateCellsWithDateAndTime(rowTarget, columnHeaderValues, "ProductSlug", sheet)
    excelFile.save(fullExcelFileName)
    excelFile.close()

def excelColumnToList(sheet, columnHeaderValues, columnName):
    rowEnd = sheet.max_row + 1
    columnNum = columnHeaderValues.index(columnName) + 1
    myList = []
    myRow = 2
    val = sheet.cell(row=myRow, column=columnNum).value
    while((val is not None) and (val != "")):
        myList.append(val)
        myRow += 1
        if(myRow < rowEnd):
            val = sheet.cell(row=myRow, column=columnNum).value
        else:
            val = None
    return myList

def modifyEndRow(sheet, columnHeaderValues, columnName, value):
    rowNum = sheet.max_row
    columnNum = columnHeaderValues.index(columnName) + 1
    sheet.cell(row=rowNum, column=columnNum).value = value

#Also contains step 11 for NRG_regression.
def stepSix(fullExcelFilename, brand):
    #print("\nIn stepSix.  This is brand: " + brand + "\n")
    #What happens when no products exist in the spreadsheet already for a given search term?  Skip them, per Bibusha.

    excelFile = openpyxl.load_workbook(fullExcelFilename)
    sheet = excelFile['NRP Product Catalog']
    lookupSheet = excelFile['Lookup']

    #Maybe correct the rowData thing later... you don't need to create that first, just use for loops and sheet.
    row_data = sheet.rows
    columnHeaderValues = []
    for cell in sheet[1]:
        columnHeaderValues.append(cell.value)

    lookupColumnHeaderValues = []
    for cell in lookupSheet[1]:
        lookupColumnHeaderValues.append(cell.value)

    stateSlugList = excelColumnToList(lookupSheet, lookupColumnHeaderValues, "States")

    #termsOfServiceTypeSet = set()
    for i in stateSlugList:
        rowNum = findARowWithThisColumnValue("StateSlug", i, columnHeaderValues, sheet, False, excelFile, True)
        if(rowNum != -1):
            copyRowToTheEndOfTheSpreadsheet(sheet, columnHeaderValues, rowNum, "Add", 0, False)
        else:
            print("A product with StateSlug " + i + " not found in the spreadsheet for " + brand + "-pos.\n")

    channelList = excelColumnToList(lookupSheet, lookupColumnHeaderValues, "Channel Slugs")

    for i in channelList:
        rowNum = findARowWithThisColumnValue("ChannelSlug", i, columnHeaderValues, sheet, False, excelFile, True)
        if(rowNum != -1):
            copyRowToTheEndOfTheSpreadsheet(sheet, columnHeaderValues, rowNum, "Add", 0, False)
        else:
            print("A product with ChannelSlug " + i + " not found in the spreadsheet for " + brand + "-pos.\n")

    termsOfServiceList = excelColumnToList(lookupSheet, lookupColumnHeaderValues, "Terms Of Service Types")

    for i in termsOfServiceList:
        rowNum = findARowWithThisColumnValue("TermsOfServiceType", i, columnHeaderValues, sheet, False, excelFile, True)
        if (rowNum != -1):
            copyRowToTheEndOfTheSpreadsheet(sheet, columnHeaderValues, rowNum, "Add", 0, False)
        else:
            print("A product with TermsOfServiceType " + i + " not found in the spreadsheet for " + brand + "-pos.\n")

    #Non-NRG_regression Retention / blank product path positive test.  Can be uncommented when I know everything else is working... might take configuration with the confirmation piece of the program.
    if((brand != "NRG_regression") and (brand != "Cirro")):
        rowNum = findARowWithThisColumnValue("ChannelSlug", "retention", columnHeaderValues, sheet, False, excelFile, True)
        if (rowNum != -1):
            copyRowToTheEndOfTheSpreadsheet(sheet, columnHeaderValues, rowNum, "Add", 0, False)
            modifyEndRow(sheet, columnHeaderValues, "ProductPath", "")
        else:
            print("A product with ChannelSlug " + "retention" + " not found in the spreadsheet for " + brand + "-pos.  Problem in the ProductPath branch.\n")

    if(brand == "NRG_regression"):
        #Add a TOU product
        rowNum = findARowWithThisColumnValue("TermsOfServiceType", "fixed", columnHeaderValues, sheet, False, excelFile, True)
        if (rowNum != -1):
            copyRowToTheEndOfTheSpreadsheet(sheet, columnHeaderValues, rowNum, "Add", 0, False)
            #We shouldn't have to set RateCategories on upload.
            #modifyEndRow(sheet, columnHeaderValues, "RateCategory", "NRGHRTUFW")
            modifyEndRow(sheet, columnHeaderValues, "PromoCode", "524")
            modifyEndRow(sheet, columnHeaderValues, "VASCode", "006")
            modifyEndRow(sheet, columnHeaderValues, "OnPeakRate", "0.20")
            modifyEndRow(sheet, columnHeaderValues, "OffPeakRate", "0.10")
            modifyEndRow(sheet, columnHeaderValues, "UtilityOnPeakRate", "0.40")
            modifyEndRow(sheet, columnHeaderValues, "UtilityOffPeakRate", "0.30")
            modifyEndRow(sheet, columnHeaderValues, "PartnerCode", "stp")
        else:
            print("A product with TermsOfServiceType " + "fixed" + " not found in the spreadsheet for " + brand + "-pos.  Problem in the \"if NRG_regression\" branch.\n")
        # Add a flat bill product.
        #rowNum = findARowWithThisColumnValue("PartnerCode", "flt", columnHeaderValues, sheet, False, excelFile, True)
        #rowNum = findARowWithThisColumnValue("TermsOfServiceType", "fixed", columnHeaderValues, sheet, False, excelFile, True)
        myBrand = "nrg_residential"
        rowNum = 2
        if (rowNum != -1):
            copyRowToTheEndOfTheSpreadsheet(sheet, columnHeaderValues, rowNum, "Add", 0, True)
            modifyEndRow(sheet, columnHeaderValues, "BrandSlug", myBrand)
        else:
            print("A product with PartnerCode " + "flt" + " not found in the spreadsheet for " + brand + "-pos.  Problem in the \"if NRG_regression\" branch.  Consider copying from a Flatbill product in a preset spreadsheet.\n")

    #if(brand == "NRG_regression"):
    #    # Add a flat bill product.
    #    #rowNum = findARowWithThisColumnValue("TermsOfServiceType", "fixed", columnHeaderValues, sheet, False, excelFile, True)
    #    rowNum = findARowWithThisColumnValue("PartnerCode", "flt", columnHeaderValues, sheet, False, excelFile, True)
    #    if (rowNum != -1):
    #        copyRowToTheEndOfTheSpreadsheet(sheet, columnHeaderValues, rowNum, "Add", 0, True)
    #        modifyEndRow(sheet, columnHeaderValues, "BrandSlug", "nrg_residential")
    #        #modifyEndRow(sheet, columnHeaderValues, "PartnerCode", "flt")
    #        #modifyEndRow(sheet, columnHeaderValues, "DailyServiceCharge", "1.64271")
    #        #modifyEndRow(sheet, columnHeaderValues, "MonthlyServiceCharge", "50")
    #        #modifyEndRow(sheet, columnHeaderValues, "PromoCode", "524")
    #        #modifyEndRow(sheet, columnHeaderValues, "VASCode", "006")
    #        #modifyEndRow(sheet, columnHeaderValues, "Rate", "0")
    #        #modifyEndRow(sheet, columnHeaderValues, "UtilityRate", "0")
    #    else:
    #        #print("A product with PartnerCode " + "flt" + " not found in the spreadsheet.  Problem in the \"if NRG_regression\" branch.  Consider copying from a Flatbill product in a preset spreadsheet.\n")

    #Here's my change block
    makeASafeChange(sheet, (sheet.max_row + 1), columnHeaderValues, excelFile)

    excelFile.save(fullExcelFilename)
    excelFile.close()

#WATCH OUT!  This isn't generic.  It's only called from the negative case test in step nine, and it uses a "False" value for complianceBool(because of the negative testing).
def findAFixedRateReadyUtilityRow(driver, sheet, workbook, columnHeaderValues):
    #forbiddenUtilities = ["Ameren", "oru", "coned"]
    forbiddenUtilities = ["Ameren"]
    myWebAddress = "http://utility.api.%s.nrgpl.us/admin/api/utility/?rate_code_strategy=stringify" % GenericSettings.getMyEnvironment().lower()
    slugXpath = "/html/body/div/div[3]/div/div/form/div/table/tbody/tr[1]/th/a"
    driver.get(myWebAddress)
    count = 1
    try:
        while(True):
            slugXpath = "/html/body/div/div[3]/div/div/form/div/table/tbody/tr[%s]/th/a" % count
            slug = driver.find_element_by_xpath(slugXpath).text
            forbiddenUtilities.append(slug)
            count = count + 1
    except:
        pass

    #xpath results from this query:
    myWebAddress = "http://utility.api.%s.nrgpl.us/admin/api/utility/?sap_bill_method=rate_ready" % GenericSettings.getMyEnvironment().lower()
    driver.get(myWebAddress)
    utilityNotFound = True
    count = 1
    try:
        while(utilityNotFound):
            slugXpath = "/html/body/div/div[3]/div/div/form/div/table/tbody/tr[%s]/th/a" % count
            slug = driver.find_element_by_xpath(slugXpath).text
            if(slug not in forbiddenUtilities):
                rowNum = findARowWithThisColumnValue("UtilitySlug", slug.lower(), columnHeaderValues, sheet, False, workbook, False)
                if(rowNum != -1):
                    return rowNum
            count = count + 1
    except:
        pass
    return -1

#Also contains step 10 for NRG_regression.
def stepNine(fullExcelFilename, brand):
    #print("\nIn stepNine.  This is brand: " + brand + "\n")
    
    driver = globals()['myDriver']
    
    workbook = openpyxl.load_workbook(fullExcelFilename)
    sheet = workbook['NRP Product Catalog']
    columnHeaderValues = []
    for cell in sheet[1]:
        columnHeaderValues.append(cell.value)

    #rowNum = findARowWithThisColumnValue("TermsOfServiceType", "fixed", columnHeaderValues, sheet, False, workbook)
    rowNum = findAFixedRateReadyUtilityRow(driver, sheet, workbook, columnHeaderValues)
    if (rowNum != -1):
        copyRowToTheEndOfTheSpreadsheet(sheet, columnHeaderValues, rowNum, "Add", 0, False)
        modifyEndRow(sheet, columnHeaderValues, "Rate", "3.9")
        modifyEndRow(sheet, columnHeaderValues, "UtilityRate", "3.90")
    else:
        print("A fixed product was not found in the spreadsheet for " + brand + "-neg.\n")

    rowNum = findARowWithThisColumnValue("TermsOfServiceType", "fixed", columnHeaderValues, sheet, False, workbook, False)
    if (rowNum != -1):
        copyRowToTheEndOfTheSpreadsheet(sheet, columnHeaderValues, rowNum, "Add", 0, False)
        modifyEndRow(sheet, columnHeaderValues, "PromoCode", "555")
    else:
        print("A product with TermsOfServiceType " + "fixed" + " not found in the spreadsheet for " + brand + "-neg.  Problem in the PromoCode branch.\n")

    rowNum = findARowWithThisColumnValue("TermsOfServiceType", "fixed", columnHeaderValues, sheet, False, workbook, False)
    if (rowNum != -1):
        copyRowToTheEndOfTheSpreadsheet(sheet, columnHeaderValues, rowNum, "Add", 0, False)
        modifyEndRow(sheet, columnHeaderValues, "EarlyCancellationFee", "")
    else:
        print("A product with TermsOfServiceType " + "fixed" + " not found in the spreadsheet for " + brand + "-neg.  Problem in the EarlyCancellationFee branch.\n")

    if((brand == "NRG_regression") or (brand == "Cirro")):
        rowNum = findARowWithThisColumnValue("ChannelSlug", "retention", columnHeaderValues, sheet, False, workbook, False)
        if (rowNum != -1):
            copyRowToTheEndOfTheSpreadsheet(sheet, columnHeaderValues, rowNum, "Add", 0, False)
            modifyEndRow(sheet, columnHeaderValues, "ProductPath", "")
        else:
            print("A product with ChannelSlug " + "retention" + " not found in the spreadsheet for " + brand + "-neg.  Problem in the ProductPath branch.\n")

    workbook.save(fullExcelFilename)
    workbook.close()

#Have to store the original input file and then swap it back into the input directory after the run... or delete the original.  I think production style2.css would involve deletion.
def mainRegressionTest(myBasePath, fileSourceList, failFileSourceList):
    #testSetup()
    globals()['usedSKUs'] = []
    deleteInputFolderContents()
    downloadSpreadsheets()
    for a in fileSourceList:
        outputFilePath = myBasePath + "ProductBuilderRegressionAutomation_SAPVersion" + os.sep + "ExcelFileProductListOutputs" + os.sep + "" + a
        outputFileBackupPath = myBasePath + "ProductBuilderRegressionAutomation_SAPVersion" + os.sep + "ExcelFileProductListOutputsBackup" + os.sep + "" + a
        filePath = myBasePath + "ProductBuilderRegressionAutomation_SAPVersion" + os.sep + "ExcelFileProductListInputs" + os.sep + "" + a
        inputBackupFilePath = myBasePath + "ProductBuilderRegressionAutomation_SAPVersion" + os.sep + "ExcelFileProductListInputsBackup" + os.sep + "" + a
        myOutputFileList = [name for name in os.listdir(outputFilePath)]
        for i in myOutputFileList:
            shutil.move(outputFilePath + os.sep + i, outputFileBackupPath + os.sep + i)
        myFileList = [name for name in os.listdir(filePath)]
        for i in myFileList:
            fullFilePath = filePath + os.sep + i
            fullInputBackupFilePath = inputBackupFilePath + os.sep + i
            #don't mirror this line below.
            #copyfile(fullFilePath, filePath + "Fail" + os.sep + "" + i)
            copyfile(fullFilePath, fullInputBackupFilePath)
            TLPSupportModule.standardizeAnExcelOrTLPFile(myBasePath, fullFilePath)
            stepSix(fullFilePath, a)
            copyfile(fullFilePath, outputFilePath + os.sep + i)

    for a in failFileSourceList:
        outputFilePath = myBasePath + "ProductBuilderRegressionAutomation_SAPVersion" + os.sep + "ExcelFileProductListOutputs" + os.sep + "" + a
        outputFileBackupPath = myBasePath + "ProductBuilderRegressionAutomation_SAPVersion" + os.sep + "ExcelFileProductListOutputsBackup" + os.sep + "" + a
        filePath = myBasePath + "ProductBuilderRegressionAutomation_SAPVersion" + os.sep + "ExcelFileProductListInputs" + os.sep + "" + a
        inputBackupFilePath = myBasePath + "ProductBuilderRegressionAutomation_SAPVersion" + os.sep + "ExcelFileProductListInputsBackup" + os.sep + "" + a
        myOutputFileList = [name for name in os.listdir(outputFilePath)]
        for i in myOutputFileList:
            shutil.move(outputFilePath + os.sep + i, outputFileBackupPath + os.sep + i)
        myFileList = [name for name in os.listdir(filePath)]
        for i in myFileList:
            fullFilePath = filePath + os.sep + i
            fullInputBackupFilePath = inputBackupFilePath + os.sep + i
            copyfile(fullFilePath, fullInputBackupFilePath)
            TLPSupportModule.standardizeAnExcelOrTLPFile(myBasePath, fullFilePath)
            myString = a.strip("Fail")
            stepNine(fullFilePath, myString)
            copyfile(fullFilePath, outputFilePath + os.sep + i)
    #throwAway = input("\nCheck brands and such.\n")
    uploadSpreadsheetsAndCheckProducts()

def checkWebPageForTextStub():
    driver = globals()['myDriver']
    myWebAddress = "http://manage.products.qa.nrgpl.us/Builder/#/upload"
    driver.get(myWebAddress)
    time.sleep(2)
    makeSureTheValidateOnlyToggleIsOff(driver)
    #checkWebPageForTextAndMaybeClick(driver, "OFF", True, False, 1, True, False)
    #checkWebPageForTextAndMaybeClick(driver, "validation error", False, False, 1, True, False)

def ProductBuilderRegression_SAPVersionTest(twoLetterEnvironment):
    throwaway = input("\nWARNING: THIS PROGRAM IS ONLY COMPATIBLE WITH A WIRED INTERNET CONNECTION.  No wifi allowed.  Please stop and make sure you are on a wired internet connection, then press any key to proceed.\n")
    GenericSettings.initializeConn()
    twoLetterEnvironment = twoLetterEnvironment.strip().upper()
    GenericSettings.setMyEnvironment(twoLetterEnvironment)
    myBasePath = GenericSettings.getTheBasePath()
    fileSourceList = ["NRG_regression", "GME", "Cirro"]
    failFileSourceList = ["NRGFail", "GMEFail", "CirroFail"]
    globals()['uniqueCounter'] = 1
    testSetup()
    mainRegressionTest(myBasePath, fileSourceList, failFileSourceList)
