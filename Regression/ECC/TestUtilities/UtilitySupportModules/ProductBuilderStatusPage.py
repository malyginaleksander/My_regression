#salting- to attach a unique identifier to something.
import atexit
import GenericSettings
import time
import openpyxl
import sys
import json
from selenium import webdriver

global myDriver

def exit_handler():
    try: globals()['myDriver'].close()
    except: pass

def testSetup():
    globals()['myDriver'] = webdriver.Chrome(r'C:\Users\drivers\chromedriver.exe')
    globals()['myDriver'].maximize_window()
    atexit.register(exit_handler)

#Assumes all "Add" and "Change" rows are at the end of the spreadsheet.
def spreadsheetProductVsAPIProductPrep(fullExcelFileName):
    excelFile = openpyxl.load_workbook(fullExcelFileName)
    sheet = excelFile['NRP Product Catalog']
    rowTarget = sheet.max_row
    columnHeaderValues = []
    for cell in sheet[1]:
        columnHeaderValues.append(cell.value)
    rowList = []
    loopContinue = True
    productNameColumnNum = columnHeaderValues.index("ProductName") + 1
    actionColumnNum = columnHeaderValues.index("Action") + 1
    while(loopContinue):
        if(sheet.cell(row=rowTarget, column=actionColumnNum).value.upper() in ("ADD", "CHANGE")):
            rowList.append(rowTarget)
            if(rowTarget <= 0):
                loopContinue = False
        else:
            loopContinue = False
        rowTarget = rowTarget - 1
    spreadsheetToAPIDict = {
    "SKU"                       : "sku",
    "CampaignId"                : "campaign_id",
    "OfferId"                   : "offer_id",
    "BrandSlug"                 : "brand_slug",
    "ChannelSlug"               : "channel",
    "SubChannel"                : "sub_channel",
    "Ranking"                   : "ranking",
    "ProductSlug"               : "product_slug",
    "ProductName"               : "product_name",
    "ProductDescription"        : "product_description",
    "ProductPath"               : "product_path",
    "ExpirationDate"            : "expiration_date",
    "StateSlug"                 : "state_slug",
    "Commodity"                 : "commodity",
    "UtilitySlug"               : "utility_slug",
    "UtilityZone"               : "utility_zone",
    "PremiseType"               : "premise_type",
    "PricingTerm"               : "pricing_term",
    "PricingEngineLockType"     : "pe_lock_type",
    "TermsOfServiceType"        : "terms_of_service_type",
    "UtilityRate"               : "utility_rate",
    "Rate"                      : "rate",
    "UtilityOnPeakRate"         : "utility_on_peak_rate",
    "OnPeakRate"                : "on_peak_rate",
    "UtilityOffPeakRate"        : "utility_off_peak_rate",
    "OffPeakRate"               : "off_peak_rate",
    "PenaltyType"               : "penalty_type",
    "EarlyCancellationFee"      : "early_cancellation_fee",
    "DailyServiceCharge"        : "daily_service_charge",
    "MonthlyServiceCharge"      : "monthly_service_charge",
    "VASCode"                   : "vas_code",
    "ProductEligibilityClass"   : "product_eligibility_class",
    "Adder"                     : "adder",            
    "PartnerCode"               : "partner_code",
    "PromoCode"                 : "promo_code",
    "MerchandiseText"           : "merchandise",
    "MerchandiseSlug"           : "merchandise_slug",
    "MerchandiseVesting"        : "merchandise_vesting",
    "SignupBonus"               : "signup_bonus",
    "SignupVesting"             : "signup_vesting",
    "OngoingValue"              : "ongoing_value",
    "OngoingFrequency"          : "ongoing_frequency",
    "PriorityLevel"             : "priority_level",
    "TermsOfServiceTemplate"    : "tos_template",
    "RequiresPriceChangeNotice" : "requires_price_change_notice",
    "RateCategory"              : "rate_category"                         
    }
    
    return [rowList, sheet, productNameColumnNum, columnHeaderValues, spreadsheetToAPIDict, excelFile]

#return True if they match, False otherwise.
#call it with the result from the function above, as shown in the line below.
#spreadsheetProductVsAPIProduct(newProductSKU, objectiveProductName, myResult[0], myResult[1], myResult[2], myResult[3], myResult[4])
def spreadsheetProductVsAPIProduct(newProductSKU, objectiveProductName, rowList, sheet, productNameColumnNum, columnHeaderValues, spreadsheetToAPIDict):
    objectiveProductNameFound = False
    objectiveProductRowNum = 0
    #print("\nThis is rowList: \n")
    #print(rowList)
    #print("\n")
    for i in rowList:
        if (sheet.cell(row=i, column=productNameColumnNum).value == objectiveProductName):
            objectiveProductRowNum = i
            objectiveProductNameFound = True
            break
    if(not objectiveProductNameFound):
        print("\nProduct Name " + objectiveProductName + " not found.\n")
        return False
    myTempString = "http://products.%s.nrgpl.us/api/v1/products/%s" % (GenericSettings.getMyEnvironment(), newProductSKU)
    jsonValid = False
    myJ = None
    loopCount = 0
    while(not jsonValid):
        myJ = GenericSettings.requestResponseChecker(myTempString, "theProductAPI")
        jsonCount = 0
        for i in myJ:
            if(jsonCount > 4):
                jsonValid = True
                break
            jsonCount = jsonCount + 1
        loopCount = loopCount + 1
        if(not jsonValid):
            if(loopCount >= 40):
                sys.exit("\nDid not get a valid return from the search: " + myTempString + " in 40 different tries in 40 minutes.  Quitting.\n")
            print("\n Presumably nothing returned from product api for search: " + myTempString + "\n")
            print("\n This is the json returned from the page search: ")
            print(json.dumps(myJ, indent=4, sort_keys=True))
            print("\n")
            print("\nSleeping for a minute, then trying the search again.  This loop will repeat up to 40 times waiting for a correct result.  Hit ctrl-C if you want to quit out instead.\n")
            time.sleep(60)
        
    myColumnNum = 5

    while(myColumnNum < (len(columnHeaderValues) + 1) ):
        cellValue = GenericSettings.nStr(sheet.cell(row=objectiveProductRowNum, column=myColumnNum).value).strip()
        columnName = columnHeaderValues[myColumnNum - 1]
        #if it's ratecategory and blank, cancel the comparison to what's in the api.... I think?
        compare = True
        if(columnName in ["ProductSlug"]):
            cellValue = cellValue.lower()
            cellValue = cellValue.replace(" ", "_")
        elif(columnName in ["RequiresPriceChangeNotice"]):
            cellValue = cellValue.lower()
            cellValue = cellValue[0].upper() + cellValue[1:]
        elif(columnName in ["EarlyCancellationFee"]):
            if(cellValue == ""):
                cellValue = ".00"
        elif(columnName in ["RateCategory"]):
            if(cellValue == ""):
                compare = False
        apiValue = "blah"
        
        if(columnName == "ExpirationDate"):
            apiDateList = GenericSettings.nStr(myJ[spreadsheetToAPIDict[columnName]]).strip().split("T")[0].split("-")
            apiValue = apiDateList[1] + "/" + apiDateList[2] + "/" + apiDateList[0]
        else:
            apiValue = GenericSettings.nStr(myJ[spreadsheetToAPIDict[columnName]]).strip()
        if(columnName == "UtilityRate"):
            cellValue = float(apiValue)
            apiValue = float(apiValue)
        if(   (columnName in ["SubChannel"]) and compare ):
            groupingString = str(myJ['groupings']).strip()
            groupingsBool = (cellValue in groupingString)
            if((cellValue != apiValue) and (not groupingsBool)):
                print("\ncolumnName " + columnName + "'s cell value " + cellValue + " not found in the product api in listing " + spreadsheetToAPIDict[columnName] + " with value " + apiValue + ".\n")
                print("\ncolumnName " + columnName + "'s cell value " + cellValue + " also not found in the product api in listing groupings with value " + groupingString + ".\n")
                print("\nobjectiveProductName is: " + objectiveProductName + "\n")
                print("\nnewProductSKU is: " + newProductSKU + "\n")
                return False
            elif( (cellValue != '') and (cellValue == apiValue) and groupingsBool ):
                print("\ncolumnName " + columnName + "'s cell value " + cellValue + " found in the product api in listing " + spreadsheetToAPIDict[columnName] + " with value " + apiValue + ".\n")
                print("\ncolumnName " + columnName + "'s cell value " + cellValue + " also found in the product api in listing groupings with value " + groupingString + ".\n")
                print("\nThe subchannel is not allowed to be found in both subchannel and groupings.\n")
                print("\nobjectiveProductName is: " + objectiveProductName + "\n")
                print("\nnewProductSKU is: " + newProductSKU + "\n")
                return False
        elif(compare):
            if(cellValue != apiValue):
                print("\ncolumnName " + columnName + "'s cell value " + str(cellValue) + " not found in the product api in listing " + spreadsheetToAPIDict[columnName] + " with value " + str(apiValue) + ".\n")
                print("\nobjectiveProductName is: " + objectiveProductName + "\n")
                print("\nnewProductSKU is: " + newProductSKU + "\n")
                return False
        myColumnNum = myColumnNum + 1
    return True

    #product api fields not used in a spreadsheet product.
    #  : "active"#: true, - no, set separately
    #  : "status"#"published" - no, set separately
    # : "backend_source"#: "northeast_retail", - no, but it's probably inferred from using the SAP- emphasis on SAP- product builder.
    # : "early_cancellation_fee_cap"#: null, - no
    # : "effective_date"#: "2019-08-20T04:22:46.140929Z", - no
    # : "line_of_business"#: "mass_market", - no
    #: "rate_subclass_code"  #: null, - no
    #: "rate_code"  #: null, - no
    # : "promo_description"#: null, - no
    # : "promo_switch_kitcode"#: null, - no
    # : "green_text"#: "0%", - no
    # : "groupings"#: [], - no
    # : "ista_product_code"#: null, - no

#From: https://stackoverflow.com/questions/25040900/scroll-down-html-table-using-selenium-and-python
def scroll_down_element(driver, element):
    try:
        driver.execute_script("arguments[0].scrollTop = 200", element)

    except Exception as e:
        print('error scrolling down web element', e)

def checkSpreadsheetVsAPIBase(driver, fullExcelFileName, fileNameOnly):
    myEnvironment = GenericSettings.getMyEnvironment().lower()
    webAddress = "http://manage.products.%s.nrgpl.us/Builder/#/status" % myEnvironment
    driver.get(webAddress)
    #As of 9/25/2019, this sleep was critical.  You can mess with how long it is, but it was needed.
    time.sleep(5)
    maxTimeInSeconds = 3600
    totalTimeUsed = 0
    fileFound = False
    count = 1
    updated_column_title_xpath = "/html/body/div/div/div/div[4]/div/div[1]/div[2]/div/div[5]/div[2]/div[1]/div[1]"
    myUpdatedColumnList = driver.find_elements_by_xpath(updated_column_title_xpath)
    for i in myUpdatedColumnList:
        i.click()
        i.click()
    time.sleep(1)
    while(not fileFound):
        fileXpath = "/html/body/div/div/div/div[4]/div/div[2]/div/div[%s]/div[2]/div[2]/div" % count
        time.sleep(2)
        myWebElement = driver.find_element_by_xpath(fileXpath)
        myFile = myWebElement.text
        #print("\nThis is myFileOnly: " + fileNameOnly + " and this is myFile: " + myFile + "\n")
        if(myFile == fileNameOnly):
            fileFound = True
            myWebElement.click()
        else:
            count = count + 1
            if(count == 6):
                if(totalTimeUsed >= maxTimeInSeconds):
                    sys.exit("\nThe NRP(SAP) product builder status page did not show the recently uploaded file " + fileNameOnly + " within an hour.  That's a failure.  Shutting down.\n")
                print("\nThe NRP(SAP) product builder status page did not show the recently uploaded file " + fileNameOnly + " yet... this is usually normal.  Sleeping for 60 seconds and refreshing the page to try again.\n")
                time.sleep(54)
                driver.refresh()
                time.sleep(5)
                myUpdatedColumnList = driver.find_elements_by_xpath(updated_column_title_xpath)
                for i in myUpdatedColumnList:
                    i.click()
                    i.click()
                totalTimeUsed = totalTimeUsed + 60
                count = 1
                time.sleep(1)  
                
    overallBatchTableXpath = "/html/body/div/div/div/div[4]/div/div[2]"
    overallBatchTable_element = driver.find_element_by_xpath(overallBatchTableXpath)
        
    #batchFirstRowXPath = "/html/body/div/div/div/div[4]/div/div[2]/div/div[1]"
    batchFirstRowLastCellXpath = "/html/body/div/div/div/div[4]/div/div[2]/div/div[1]/div[5]/div[2]/div"
    myFirstRow = driver.find_element_by_xpath(batchFirstRowLastCellXpath)
    time.sleep(2)
    #myFirstRow.click()
    #myResult = spreadsheetProductVsAPIProductPrep("C:\\Users\\mhissong\\Desktop\\Regression\\Regression\\ECC\\TestUtilities\\ProductBuilderRegressionAutomation_SAPVersion\\ExcelFileProductListInputs\\Cirro\\Products_9_10_2019 9_12_03 AM_C.xlsx")
    myResult = spreadsheetProductVsAPIProductPrep(fullExcelFileName)
    count = 1
    #I'll need to figure out what to do if I have to scroll... I might also want to just get this information from product_api database.  IDK.
    for i in myResult[0]:
        time.sleep(1)
        #print("\nThis is count: " + str(count) + "\n")
        skuXpath = "/html/body/div/div/div/div[5]/div[2]/div[2]/div/div[%s]/div[4]/div[2]/div/span" % str(count)
        newProductSKUElement = driver.find_element_by_xpath(skuXpath)
        newProductSKUElement.click()
        newProductSKU = newProductSKUElement.text
        objProdNameXpath = "/html/body/div/div/div/div[5]/table/tbody/tr[10]/td"
        objProdNameElement = driver.find_element_by_xpath(objProdNameXpath)
        #objectiveProductName = objProdNameElement.text.strip()
        objectiveProductName = objProdNameElement.text
        if(not spreadsheetProductVsAPIProduct(newProductSKU, objectiveProductName, myResult[0], myResult[1], myResult[2], myResult[3], myResult[4])):
            print("\nThe spreadsheet and api information do not match.\n")
            myResult[5].close()
            return False
        if(count % 6 == 0):
            scroll_down_element(driver, overallBatchTable_element)
        count = count + 1
    myResult[5].close()
    return True

#if __name__ == '__main__':
#    #sys.settrace(trace_calls)
#    main()
